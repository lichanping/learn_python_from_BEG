import csv
import os
import re
import time
from datetime import datetime

import crawlertool as tool
import pandas as pd
from Selenium4R import Chrome
from bs4 import BeautifulSoup
from openpyxl import load_workbook

OUTPUT_DIR_NAME = r"E:\Temp"
NAME_MAP_URLS = [
    ['MK-少女情怀', 'https://www.huya.com/22819541'],
    ['快乐从和辣妹一起蹦迪开始', 'https://www.huya.com/17055318'],
    ['鱼塘_虎牙只有一个象_那就是大象', 'https://www.huya.com/78315'],
    ['年华：御用歌手公子轩', 'https://www.huya.com/21374754']
]
INDEX = 0
COUNT_NUMBER = 60 * 10


class SpiderHuyaLiveBarrage(tool.abc.LoopSpider):
    """虎牙直播弹幕爬虫"""

    def __init__(self, driver, live_url, interval, count_number=10):
        super().__init__(interval)
        self.driver = driver
        self.text_subscribe = 0

        # 访问目标虎牙主播的直播间
        self.driver.get(live_url)

        # 读取直播间订阅数量
        label_subscribe = self.driver.find_element_by_xpath('//*[@id="activityCount"]')
        if label_subscribe:
            self.text_subscribe = label_subscribe.text

        # 等待页面渲染完成
        time.sleep(10)

        self.data_id_max = 0
        self.count_number = count_number

    def get_noble_level(self, text):
        if isinstance(text, list):
            text = ' '.join(text)
        assert isinstance(text, str), 'nobel class type incorrect'
        pattern = r'box-noble-level-(\d+)'
        match = re.search(pattern, text)
        if match:
            noble_level_id = match.group(1)
            return noble_level_id

    def running(self):
        label_html = self.driver.find_element_by_id("chat-room__list").get_attribute("innerHTML")
        bs = BeautifulSoup(label_html, "lxml")  # 将网页内容解析为Soup对象

        barrage_list = []
        for label in bs.select("li"):
            data_id = int(label["data-id"])  # 提取:弹幕ID

            if data_id <= self.data_id_max:  # 依据弹幕的ID判断弹幕是否还未抓取
                if data_id > self.data_id_max - 101:
                    continue

            print("弹幕ID：{}".format(data_id))
            self.data_id_max = data_id

            barrage_info = {
                "bid": str(data_id),  # 弹幕ID
                "type": "",  # 弹幕所属类型
                "user_name": "",  # 弹幕发布者名称
                "user_noble": "",  # 弹幕发布者贵族等级
                "content": "",  # 弹幕内容
                "gift_name": "",  # 礼物名称
                "gift_num": "",  # 礼物数量
                "other": "",  # 其他信息
                "fetch_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 弹幕采集时间
            }

            category = str(label.select_one("li > div")["class"])  # 提取:弹幕类型
            if "msg-smog" in category:  # 处理smog类型弹幕(普通弹幕)
                barrage_info["type"] = "SG(smog类型弹幕)"
                barrage_info["user_name"] = label.select_one("li > div > span:nth-child(1)").text
                barrage_info["content"] = label.select_one("li > div > span:nth-child(3)").text
            elif "msg-normal" in category:  # 处理普通类型弹幕(普通弹幕)
                barrage_info["type"] = "NM(普通弹幕)"
                barrage_info["user_name"] = label.select_one("li > div > span:nth-child(2)").text
                barrage_info["content"] = label.select_one("li > div > span:nth-child(5)").text
            elif "msg-nobleEnter" in category:  # 处理nobleEnter类型弹幕(贵族进入弹幕)
                barrage_info["type"] = "NE(贵族进入)"
                barrage_info["user_name"] = label.select_one("li > div > div > p > span:nth-child(1)").text
                user_noble_class = label.select_one("li > div > div")["class"]
                barrage_info["user_noble"] = str(self.get_noble_level(user_noble_class))
                barrage_info["content"] = "驾临直播间"
            elif "msg-nobleSpeak" in category:  # 处理nobleSpeak类型弹幕(贵族发言)
                barrage_info["type"] = "NS(贵族发言)"
                barrage_info["user_name"] = label.select_one("li > div > span").text
                user_noble_class = label.select_one("li > div")["class"]
                barrage_info["user_noble"] = str(self.get_noble_level(user_noble_class))
                barrage_info["content"] = label.select_one("li > div > span:nth-child(5)").text
            elif "tit-h-send" in category:  # 处理send类型提示(礼物赠送提示)
                barrage_info["type"] = "SD(礼物赠送)"
                barrage_info["user_name"] = label.select_one("li > div > span:nth-child(1)").text
                barrage_info["gift_name"] = label.select_one("li > div > span:nth-child(3) > img")["alt"]
                barrage_info["gift_num"] = str(label.select_one("li > div > span:nth-child(4)").text)
            elif "msg-onTVLottery" in category:
                barrage_info["type"] = "TV"
                barrage_info["user_name"] = label.select_one("li > div > span:nth-child(2)").text
                barrage_info["content"] = label.select_one("li > div > div > span").text
            elif "msg-auditorSys" in category:  # 处理msg-auditorSys类型提示(系统提示)
                barrage_info["type"] = "AS(msg-auditorSys类型提示)"
                barrage_info["other"] = label.text
            elif "msg-sys" in category:  # 处理msg-sys类型提示(系统提示)
                barrage_info["type"] = "SY(msg-sys类型提示)"
                barrage_info["other"] = label.text
            else:  # 处理其他类型
                barrage_info.update(type="OT", other="弹幕名称" + category)

            raw = [
                barrage_info['bid'], barrage_info['type'], barrage_info['user_name'],
                barrage_info['user_noble'], barrage_info['content'], barrage_info['gift_name'],
                barrage_info['gift_num'], barrage_info['other'], barrage_info['fetch_time'],
            ]
            barrage_list.append(raw)

        if barrage_list:
            self.write(barrage_list)
        # print("爬虫循环：{}".format(self.num))
        if self.num >= self.count_number:
            self.stop()


# ------------------- 单元测试 -------------------
if __name__ == "__main__":
    class MySpider(SpiderHuyaLiveBarrage):
        """重写SpiderBilibiliLiveBarrage类"""

        def __init__(self, driver, live_url, count_number, file_name):
            super().__init__(driver=driver, live_url=live_url, interval=1, count_number=count_number)
            self.file_name = file_name
            self.file_path = os.path.join(OUTPUT_DIR_NAME,
                                          r'{}.xlsx'.format(self.file_name))

        def write(self, data):
            df = pd.DataFrame(data=data, columns=['弹幕ID',
                                                  '弹幕所属类型',
                                                  '弹幕发布者名称',
                                                  '贵族等级',
                                                  '弹幕内容',
                                                  '礼物名称',
                                                  '礼物数量',
                                                  '其他信息',
                                                  '弹幕采集时间'])
            if os.path.isfile(self.file_path):
                sheetname = 'Sheet1'
                writer = pd.ExcelWriter(self.file_path, engine='xlsxwriter')
                load_workbook(self.file_path)
                writer.book = load_workbook(self.file_path)
                startrow = writer.book[sheetname].max_row
                df.to_excel(writer, sheet_name=sheetname, startrow=startrow, index=False,
                            header=False)

            else:
                df.to_excel(self.file_path, index=False, header=True)

        def stop(self):
            subscribe_count = str(self.text_subscribe)
            print("Total Subscribe Count: {}".format(subscribe_count))

            super(MySpider, self).stop()


    driver = Chrome(cache_path=OUTPUT_DIR_NAME)  # 打开Chrome浏览器
    spider = MySpider(driver=driver, live_url=NAME_MAP_URLS[INDEX][1],
                      count_number=COUNT_NUMBER, file_name=NAME_MAP_URLS[INDEX][0])
    spider.start()
